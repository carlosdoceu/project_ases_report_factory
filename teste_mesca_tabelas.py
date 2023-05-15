from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd

# Criar um novo arquivo Excel
workbook = Workbook()
sheet = workbook.active

# Preencher a tabela com alguns dados
sheet['A1'] = 'Valor1'
sheet['A2'] = 'Valor2'
sheet['A3'] = 'Valor3'
sheet['B1'] = 'Valor4'
sheet['B2'] = 'Valor5'
sheet['B3'] = 'Valor6'

# Mesclar células
sheet.merge_cells('C1:D1')
sheet.merge_cells('C2:D2')
sheet.merge_cells('C3:D3')

# Ajustar a alinhamento das células mescladas
for row in sheet['C1':'D3']:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Salvar o arquivo Excel
workbook.save('arquivo.xlsx')

# Converter para CSV usando pandas
df = pd.read_excel('arquivo.xlsx')
df.to_csv('arquivo.csv', index=False)
